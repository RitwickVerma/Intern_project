from openpyxl import *

class ExcelSheet:
    def __init__(self,excelsheet):                                       
        self.excelsheet=excelsheet
        self.set_current_row()

    #gets current row
    def get_current_row(self):
        return self.current_row
    
    #sets current row
    def set_current_row(self,row=1):
        self.current_row=row

    #set title of the worksheet
    def set_title(self,name):
        self.excelsheet.title=name

    #get title of the worksheet
    def get_title(self):
        return self.excelsheet.title

    #prints a complete row into excel sheet and pointer(current_row) is incremented (1 based indexing)
    #If row is specified, data will be overwritten in that row
    def print_row(self,datalist,row=None):
        if(row==None):
            row=self.get_current_row()
        
        for i in range(len(datalist)):
            self.excelsheet.cell(row=row,column=i+1).value=datalist[i]
        
        if(row==self.get_current_row()):
            self.current_row+=1
        elif(row>self.current_row):
            self.current_row=row
    


class ExcelFile:
    def __init__(self):
        self.excelfile=Workbook()
        excelsheet=ExcelSheet(self.excelfile.active)  #by default every excel file is created with atleast one worksheet
        self.excelsheets=[]
        self.excelsheets.append(excelsheet)
        self.current_active_sheet_index=0

    #set and/or get active sheet from all sheets by index
    def active_excelsheet(self,index=None):
        if(index!=None):
            self.excelfile.active=index
            self.current_active_sheet_index=index
    
        return self.excelsheets[self.current_active_sheet_index]
        
    #get index of active worksheet
    def get_active_excelsheet_index(self):
        return self.current_active_sheet_index
    
    #adding new worksheet. Newly added worksheet is SET ACTIVE by default
    def add_excelsheet(self,name,index=None):
        if(index==None):  
            worksheet=self.excelfile.create_sheet(name)
            excelsheet=ExcelSheet(worksheet)
            self.excelsheets.append(excelsheet)
            self.active_excelsheet(len(self.excelsheets)-1)
        else:
            worksheet=self.excelfile.create_sheet(name,index)
            excelsheet=ExcelSheet(worksheet)
            self.excelsheets.insert(excelsheet,index)
            self.active_worksheet(index)

    #can be used temporarily if function isn't available in wrapper. This gives complete access to openpyxl
    def get_openpyxl_workbook_object(self):
        return self.excelfile

    #save the excel file
    def save(self,name='data.xlsx'):
        self.excelfile.save(name)

    

#returns two lists of all fields and all corresponding values 
def get_fields_values(model,include_id=False):
    fieldlist=[]
    valuelist=[]
    for key,value in model.items():
        if(key!='id'):
            fieldlist.append(key)
            valuelist.append(value)
    return fieldlist,valuelist


