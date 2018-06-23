from copy import deepcopy
import openpyxl
from openpyxl.drawing.image import Image
import importlib
from django.conf import settings
from core.resources import is_blank_or_none
from PIL import Image as PilImage
import os
class Excel():
    __cell = {'column': None, 'row': None, 'value': '', 'valuetype': 'str'}
    def __init__(self, template_name, sheet_names=None):
        self.template_name = template_name
        self.sheet_names = sheet_names
        self.worksheet = []
        for sheet in sheet_names:
            self.worksheet.append([])
        self.wb = openpyxl.load_workbook(str(self.template_name))
    def add_cell(self, index, column, row, value=__cell['valuetype'], value_type=__cell['valuetype']):
        cell = deepcopy(self.__cell)
        cell['column'] = column
        cell['row'] = row
        cell['value'] = value
        cell['value_type'] = value_type
        self.worksheet[index].append(cell)
    def __convert_to_type(self, element, type_):
        if is_blank_or_none(element):
            element = ''
        converted_element = str(element).strip('\"').replace('\\n',"\n")
        try:
            # Check if it's a builtin type
            module = importlib.import_module('__builtin__')
            cls = getattr(module, type_)
            converted_element = cls(converted_element)
            # print "converted_element",converted_element
        except:
            pass
        return converted_element
    def export_xlsx(self):
        for index, sheet_name in enumerate(self.sheet_names):
            ws = self.wb.get_sheet_by_name(str(sheet_name))
            for cell in self.worksheet[index]:
                cell_location = str(cell['column'])+str(cell['row'])
                cell_value = cell['value']
                if cell['value_type'] == "url" and cell_value is not None:
                    cell_value = os.path.abspath("{}/{}".format(settings.MEDIA_ROOT, cell_value))
                    # cell_value = PilImage.open(cell_value)
                    # # Reduced the size of the image by reducing its quality
                    # result = cell_value.convert("P")
                    img = Image(cell_value)
                    img.anchor(ws[cell_location], anchortype='oneCell')
                    ws.add_image(img)
                elif (cell['value_type'] == "multiple" and cell_value is not None):
                    import ast
                    try:
                        cell_value = ast.literal_eval(str(cell_value))
                        ws[cell_location] = ','.join(str(value) for value in cell_value)
                    except:
                        pass
                else:
                    ws[cell_location] = self.__convert_to_type(cell_value, cell['value_type'])
        return self.wb
